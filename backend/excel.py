"""Excel (.xlsx) report generation.

Unlike PDF/reportlab, openpyxl needs no manual bidi handling for Hebrew -
setting sheet_view.rightToLeft=True makes Excel display column A on the
right and flip cell text direction on its own.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from months import MONTH_NAMES
from pdf import FIELD_HEADERS

HEADER_FILL = PatternFill(start_color="F0EEFB", end_color="F0EEFB", fill_type="solid")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
AMOUNT_FORMAT = '#,##0" ש״ח"'  # ...ש"ח

# column widths, aligned with FIELD_HEADERS order
FIELD_WIDTHS = [12, 14, 14, 12, 10, 10, 10, 10, 12, 10, 10, 14]


def _yesno(value):
    return "כן" if value else "לא"


def _record_row(record):
    return [
        record.get("study_hours"),
        record.get("excluded_hours"),
        record.get("attendance_amount"),
        _yesno(record.get("with_american")),
        _yesno(record.get("emuna")),
        _yesno(record.get("tanach")),
        _yesno(record.get("ktiva")),
        _yesno(record.get("gemara_bekiut")),
        _yesno(record.get("review_test")),
        _yesno(record.get("enrichment")),
        _yesno(record.get("reserve_duty")),
        record.get("total_amount"),
    ]


AMOUNT_COLUMNS = {4, 13}  # sheet column numbers holding attendance_amount / total_amount


def _write_sheet(ws, title, row_label_header, row_label_width, rows):
    ws.sheet_view.rightToLeft = True

    ncols = 1 + len(FIELD_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    headers = [row_label_header] + FIELD_HEADERS
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (label, record) in enumerate(rows, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        for col_offset, value in enumerate(_record_row(record), start=2):
            cell = ws.cell(row=row_idx, column=col_offset, value=value)
            if col_offset in AMOUNT_COLUMNS:
                cell.number_format = AMOUNT_FORMAT
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions[get_column_letter(1)].width = row_label_width
    for col, width in enumerate(FIELD_WIDTHS, start=2):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def build_month_report_xlsx(year, month, avreichim_records):
    """avreichim_records: list of (avrech_name, record_dict)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח חודשי"

    month_name = MONTH_NAMES[month - 1]
    title = f"דוח מלגות לכל האברכים - {month_name} {year}"
    rows = [(name, record) for name, record in avreichim_records]

    _write_sheet(ws, title, "שם אברך", 20, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_avrech_report_xlsx(avrech_name, year, records_by_month):
    """records_by_month: list of 12 record dicts, index 0 = January"""
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח שנתי"

    title = f"דוח מלגות - {avrech_name} - שנת {year}"
    rows = [(MONTH_NAMES[i], records_by_month[i]) for i in range(12)]

    _write_sheet(ws, title, "חודש", 12, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
